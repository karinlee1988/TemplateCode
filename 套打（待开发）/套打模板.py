#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2022/2/8 9:44
# @Author : karinlee
# @FileName : 套打模板.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976

"""
用于套打用。将一个excel模板表格多行的内容套打到excel表或word文档中，模板表格每行生成一份套打的excel表或word文档

20220306 开发中
"""
import os
import time


import openpyxl
from docx import Document

# def replace_text(doc:[Document],old_text:str, new_text:str) -> None:
#     """
#     将某个文档的某个词进行替换
#
#     :param doc:   要处理的文件
#     :type doc: Document
#
#     :param old_text:  旧的词
#     :type old_text: str
#
#     :param new_text: 新的词
#     :type new_text: str
#
#     :return: None
#     """
#     for p in doc.paragraphs:
#         if old_text in p.text:
#             inline = p.runs
#             for i in inline:
#                 if old_text in i.text:
#                     text = i.text.replace(old_text, new_text)
#                     i.text = text
#
#
# def main():
#     """
#     入口
#     :return:
#     """
#
#     wb = openpyxl.load_workbook("20211112_向各企业寄送提供稳岗补贴银行账户通知（570人）")
#     ws = wb.active
#     # -------------------
#     # 这里修改开始编号->结束编号
#     tag_begin = 501
#     tag_end = 570
#     #-------------------
#     for row in range(tag_begin+2,tag_end+3):
#         document_temp = Document("关于提供单位银行开户资料的通知（模板）.docx")
#         tag = str(ws.cell(row=row,column=1).value)
#         sjr = ws.cell(row=row,column=2).value
#         phone = ws.cell(row=row,column=3).value
#         address = ws.cell(row=row,column=4).value
#         company = ws.cell(row=row,column=8).value
#
#         replace_text(document_temp,"{tag}",tag)
#         replace_text(document_temp, "{sjr}", sjr)
#         replace_text(document_temp, "{address}", address)
#         replace_text(document_temp, "{phone}", phone)
#         replace_text(document_temp, "{company}", company)
#         document_temp.save('通知\\'+tag+".docx")
#         os.startfile('通知\\'+tag+".docx",'print')
#         time.sleep(5)


class TemPrint(object):

    def __init__(self,template_path,source_path,begin_count,end_count,sleeper):

        self.source_path = source_path
        self.template_path = template_path
        self.begin_count = int(begin_count)
        self.end_count = int(end_count)
        self.template_sheet_index = 0
        self.source_sheet_index = 0
        self.workbook_template = openpyxl.load_workbook(self.template_path)
        self.workbook_source = openpyxl.load_workbook(self.source_path)
        self.worksheet_template = self.workbook_template[self.workbook_template.sheetnames[self.template_sheet_index]]
        self.worksheet_source = self.workbook_source[self.workbook_source.sheetnames[self.source_sheet_index]]
        self.sleeper = int(sleeper)


    def fill_and_print(self):


        for row in range(self.begin_count,self.end_count+1):
            name = self.worksheet_source.cell(row=row, column=6).value
            sfz = self.worksheet_source.cell(row=row,column=7).value
            if self.worksheet_source.cell(row=row, column=3).value:
                self.worksheet_template['D4'].value = self.worksheet_source.cell(row=row,column=6).value
                self.worksheet_template['D5'].value = self.worksheet_source.cell(row=row, column=7).value
                self.worksheet_template['D6'].value = self.worksheet_source.cell(row=row, column=3).value
                self.worksheet_template['N6'].value = self.worksheet_source.cell(row=row, column=2).value
                self.worksheet_template['B23'].value = self.worksheet_source.cell(row=row, column=8).value
                self.worksheet_template['D23'].value = self.worksheet_source.cell(row=row, column=9).value

                self.worksheet_template['A1'].value = row

                self.workbook_template.save("printtable\\" + "个人社会保险补缴申请表--" + name + '.xlsx')
                os.startfile("printtable\\"+"个人社会保险补缴申请表--"+ name +'.xlsx','print')
                print(f">>> 行号：{row},姓名： {name} ，身份证号码：{sfz} 已处理完成！<<<")
                time.sleep(self.sleeper)      #  延时sleeper秒后继续下一人员打印

class TemprintGUI(object):
    """
    可视化界面
    """

    def __init__(self):
        # 设置pysimplegui主题，不设置的话就用默认主题
        # sg.ChangeLookAndFeel('Purple')
        # 定义2个常量，供下面的layout直接调用，就不用一个个元素来调字体了
        # 字体和字体大小
        self.FONT = ("微软雅黑", 12)
        # 可视化界面上元素的大小
        self.SIZE = (18, 1)
        # 界面布局
        self.layout = [
            # sg.Image()插入图片，支持gif和png
            [sg.Image(filename="peppa.png", pad=(100,0))],
            # sg.Text()显示文本
            [sg.Text('', font=self.FONT, size=self.SIZE)],
            # sg.Input()是输入框
            [sg.Text('请选择模板：', font=self.FONT, size=(30, 1))],
            [sg.Input('  ', key="_FILETEMP_", readonly=True, size=(36, 1), font=self.FONT),
             sg.FileBrowse(button_text='选择模板', size=(10, 1), font=self.FONT)],
            [sg.Text('请选择数据来源：', font=self.FONT, size=(30, 1))],
            [sg.Input('  ', key="_FILESOURCE_", readonly=True, size=(36, 1), font=self.FONT),
             sg.FileBrowse(button_text='选择数据', size=(10, 1), font=self.FONT)],
            [sg.Text(' 请输入起始行号：', font=self.FONT, size=self.SIZE),
             sg.Input(key='_BEGIN_', font=self.FONT, size=(10, 1))],
            [sg.Text(' 请输入结束行号:  ', font=self.FONT, size=self.SIZE),
             sg.Input(key='_END_', font=self.FONT, size=(10, 1))],
            [sg.Text(' 请输入间隔时间（秒）:  ', font=self.FONT, size=self.SIZE),
             sg.Input(key='_TIME_', font=self.FONT, size=(10, 1))],
            [sg.Text('')],
            [sg.Btn('开始套打', key='_SUMMIT_', font=("微软雅黑", 16), size=(20, 1)),
             sg.Input(key='_STATUS_', readonly=True, size=(10, 1), font=self.FONT)],
            [sg.Text('')],
            [sg.Output(size=(55, 5),font=("微软雅黑", 10),background_color='light gray')]
        ]
        # 创建窗口，引入布局，并进行初始化
        # 创建时，必须要有一个名称，这个名称会显示在窗口上
        self.window = sg.Window('个人社会保险补缴申请表套打by英德关系李加林', layout=self.layout, finalize=True)


    # 窗口持久化
    def run(self):
        # 创建一个事件循环，否则窗口运行一次就会被关闭
        while True:
            # 监控窗口情况
            event, value = self.window.Read()
            # 当获取到事件时，处理逻辑（按钮绑定事件，点击按钮即触发事件）
            if event == '_SUMMIT_':
                temp = value['_FILETEMP_']
                source = value['_FILESOURCE_']
                begin = value['_BEGIN_']
                end = value['_END_']
                sleep = value['_TIME_']
                table = TemPrint(temp,source,begin,end,sleep)
                table.fill_and_print()



                # 调用函数拆分表格

                # 函数完成后返回处理完成标志到窗口界面上
                self.window.Element("_STATUS_").Update("处理完成！")
            # 如果事件的值为 None，表示点击了右上角的关闭按钮，则会退出窗口循环
            if event is None:
                break
        self.window.close()

if __name__ == '__main__':

    tablegui = TemprintGUI()
    tablegui.run()

