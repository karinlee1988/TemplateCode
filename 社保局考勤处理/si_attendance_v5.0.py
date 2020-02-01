#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 20200126
# @Author  : karinlee
# @FileName: si_attendance_v5.0.py
# @Software: PyCharm
# @Blog    ：https://blog.csdn.net/weixin_43972976

"""
本模块用于处理社保局每月考勤的excel表格
v5.0  GUI界面

待处理表结构：
 ['序号', '姓名', '日期', '对应时段', '上班时间', '下班时间', '签到时间', '签到情况']
 ['69', '李伟斌', '2019/1/14', '上午', '08:30', '12:00', '08:19', '']

 ...

"""

import openpyxl
import tkinter as tk
import tkinter.filedialog

class SiAttendance(object):

    def __init__(self,full_path,month):
        """
        构造函数 定义实例的workbook对象， worksheet对象，考勤日期，财保和平安的人员名单list

        :param full_path: 考勤原始文件的全文件名（路径+文件名）
        :type  full_path: str

        :param month: 考勤月份
        :type  month: str

        """
        self.full_path = full_path
        self.month = month
        self.wb = openpyxl.load_workbook(self.full_path)
        self.ws_main = self.wb.active
        self._caibao = ['姓名', '李伟斌', '张伦', '张小谊', '吴雪红', '林武英']  # 财保人员
        self._pingan = ['姓名', '曾金龙', '邓翠媚', '张正文', '邱锦鹏', '钟贤威']  # 平安人员

    def change_worksheet_title(self):
        """更改sheet页的名称"""
        self.ws_main.title = u'社保局考勤明细' + self.month

    def deal_si_attendance(self):
        """初始化处理考勤明细表（全部人员）"""
        self.ws_main['H1'].value = u"签到情况"
        for row in range(1, self.ws_main.max_row + 1):
            # 将 H列的 True 更改为 未按规定时间签到
            if self.ws_main.cell(row=row, column=8).value == 'True':
                self.ws_main.cell(row=row, column=8).value = u"未按规定时间签到"

    def deal_department_attendance(self,worksheet,people_list):
        """
        进行各部门的人员考勤筛选处理

        :param worksheet: 新建的工作表用于存放筛选后的人员考勤
        :type worksheet: class worksheet

        :param people_list:  需要处理的人员名单
        :type people_list:  list

        :return:  None

        """
        list = []
        list_row = []
        for row in range(1, self.ws_main.max_row + 1):
            if self.ws_main.cell(row=row, column=2).value in people_list: # 遍历所有行 对姓名所在列的姓名与人员列表比较
                for cell in self.ws_main[row]:  # 对当前行遍历所有单元格
                    list_row.append(cell.value) # list_row是临时的1维列表 在遍历单元格获得每个单元格的值后写入列表 从而存储当前行的数据
                list.append(list_row)  # list是二维列表 里面的每个元素都是1个list_row
                list_row = []   # 重新初始化list_row列表
        for row in list:  # 对于2维列表的每个元素（每个元素就是每个1维列表 这些1维列表就等于excel表中1行的数据
            worksheet.append(row)  # 二维列表可以直接使用worksheet.append()方法写入工作表中

    def save_as(self):
        """
        处理后的数据进行保存

        """
        sheetnames = self.wb.sheetnames
        for i in range(len(sheetnames)-1,-1,-1):
            if sheetnames[i] == 'Sheet1' or sheetnames[i] == 'Sheet2' or sheetnames[i] == 'Sheet3':
                sheetnames.remove(sheetnames[i])
        for name in sheetnames:
            ws = self.wb[name]
            # 创建新的Excel
            wb_new = openpyxl.Workbook()
            # 获取当前sheet
            ws_new = wb_new.active
            # 两个for循环遍历整个excel的单元格内容
            for i, row in enumerate(ws.iter_rows()):
                for j, cell in enumerate(row):
                    # 写入新Excel
                    ws_new.cell(row=i + 1, column=j + 1, value=cell.value)
                    # 设置新Sheet的名称
                    ws_new.title = name
            wb_new.save(name + '.xlsx')

    def dealing(self):
        """
        进行考勤报表整个流程的处理

        """
        ws_caibao = self.wb.create_sheet(u'李伟斌等5人考勤' + self.month)
        ws_pingan = self.wb.create_sheet(u'曾金龙等5人考勤' + self.month)
        self.change_worksheet_title()
        self.deal_si_attendance()
        self.deal_department_attendance(ws_caibao,self._caibao)
        self.deal_department_attendance(ws_pingan, self._pingan)

        self.save_as()

# wb.save('kaoqin_done.xlsx')

# if __name__ == '__main__':
#
#     si = SiAttendance('201912.xlsx','201912')
#     si.dealing()

class SiGui(object):
    """
    tkinter GUI处理考勤
    """

    def __init__(self):
        """
        创建界面
        """
        # 新建窗口
        self.master = tk.Tk()
        self.photo = tk.PhotoImage(file="si2.gif")
        # self.path 用于存放选择的文件路径
        # self.flag暂时没用
        # self.v1 self.v2接收用户参数
        # 注意！这些都是tk.StringVar()对象，不是str，其他地方要用的话要用get()方法获取str
        self.path = tk.StringVar()
        self.flag = tk.StringVar()
        self.v1 = tk.StringVar()
        self.v2 = tk.StringVar()
        # 窗口图片横幅
        self.img_pack()
        # 主界面布局
        self.window()
        # 框架布局
        self.frame()
        # 保持运行
        self.master.mainloop()

    def img_pack(self):
        """
        窗口图片横幅
        :return:
        :rtype:
        """
        # 图片贴上去
        imgLable = tk.Label(self.master, image=self.photo)
        imgLable.pack()

    def window(self):
        """
        主界面布局设置
        :return:
        :rtype:
        """
        self.master.geometry("800x600+600+100")
        self.master.title("社保局考勤处理by李加林v5.0")
        # 图片贴上去
        # self.photo = tk.PhotoImage(file="xzpq.gif")
        # imgLable = tk.Label(self.master, image=self.photo)
        # imgLable.pack()
        # 把标题贴上去
        tk.Label(self.master,text="社保局考勤处理",font=("黑体",18)).pack()
        tk.Label(self.master,text="(自动对考勤系统导出的原始数据进行处理并分离出财保和平安人员)",font=("黑体",14)).pack()

    def frame(self):
        """
        框架布局设置
        :return:
        :rtype:
        """
        # 框架贴上去，再在框架里添加Lable，Entry，Button等控件
        frame1 = tk.Frame(self.master)
        frame1.pack()
        # 输入框，标记，按键
        tk.Label(frame1, text="目标路径:", font=("黑体", 14)).grid(row=1, column=0)
        tk.Entry(frame1, textvariable=self.path, width=50).grid(row=1, column=1)
        tk.Button(frame1, text="路径选择", command=self.select_path, font=("黑体", 16)).grid(row=1, column=2,padx=10)
        tk.Label(frame1, text="请输入考勤日期：", font=("黑体", 14)).grid(row=2, column=0)
        tk.Entry(frame1, textvariable=self.v1).grid(row=2, column=1)
        # tk.Label(frame1, text="参数2:", font=("黑体", 16)).grid(row=3, column=0)
        # tk.Entry(frame1, textvariable=self.v2).grid(row=3, column=1)
        # 按这个按钮执行主程序
        tk.Button(frame1, text="开始处理", command=self.main, font=("黑体", 14)).grid(row=4, column=1,pady=33)
        tk.Entry(frame1, textvariable=self.flag,state="readonly").grid(row=5, column=1)

    def select_path(self):
        """
        选择文件，并获取文件的绝对路径
        :return:
        :rtype:
        """
        # 选择文件，path_select变量接收文件地址
        path_select = tkinter.filedialog.askopenfilename()
        # 通过replace函数替换绝对文件地址中的/来使文件可被程序读取
        # 注意：\\转义后为\，所以\\\\转义后为\\
        path_select1 = path_select.replace("/", "\\\\")
        # self.path设置path_select的值
        self.path.set(path_select1)

    def main(self):
        """
        这个是主程序
        :return:
        :rtype:
        """
        filepath = self.path.get()
        date = self.v1.get()
        si = SiAttendance(filepath,date)
        si.dealing()
        # 标志设置为处理完成
        self.flag.set("处理完成！")

if __name__ == '__main__':
    app = SiGui()
