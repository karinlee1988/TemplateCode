#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2020/2/14 15:59
# @Author : karinlee
# @FileName : vlookup_gui_v3.0.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976

import os
import openpyxl
import tkinter as tk
import tkinter.filedialog
from openpyxl.utils import get_column_letter, column_index_from_string

def vlookup(
        wb_template,
        ws_template_index,
        template_key,
        template_value,
        wb_source,
        ws_source_index,
        source_key,
        source_value,
        line
        ):
    """
    对2个不同的工作薄执行vlookup操作

    模板工作薄：需要写入数据的工作薄
    数据工作薄：根据模板工作薄提供的条件（列）在数据工作薄中查找，提供数据来源的工作薄
    注意！函数执行完后，只对wb_template对象进行了数据写入。在函数外部还需wb_template.save("filename.xlsx"),vlookup后的数据才能保存为excel表。

    :param:
        'wb_template': 模板工作薄对象
        'ws_template_index': 需要处理的模板工作表索引号
        'template_key': 模板工作表key所在列号
        'template_value': 模板工作表value需要填写的列号
        'wb_source': 数据工作薄对象
        'ws_source_index': 需要处理的数据工作表索引号
        'source_key':  数据工作表key所在列号
        'source_value': 数据工作表value所在列号
        'line' :从第几行开始vlookup

    :type:
        'wb_template': class workbook
        'ws_template_index': int
        'template_key': str
        'template_value': str
        'wb_source': class workbook
        'ws_source_index': int
        'source_key':  str
        'source_value': str
        'line' :int

    :return: None

    """
    # 获取数据工作表对象
    ws_source = wb_source[wb_source.sheetnames[ws_source_index]]
    # 获取模板工作表对象
    ws_template = wb_template[wb_template.sheetnames[ws_template_index]]
    # 获取数据工作表的查找列和数据列，分别生成2个元组
    source_key_tuple = ws_source[source_key]
    source_value_tuple = ws_source[source_value]
    # 创建2个列表，遍历元组，将元组中每个单元格的值添加到列表中
    list_key =[]
    list_value = []
    for cell in source_key_tuple:
        list_key.append(cell.value)
    for cell in source_value_tuple:
        list_value.append(cell.value)
    # 通过数据工作表的key列和value列，创建好需要进行vlookup的字典
    dic = dict(zip(list_key,list_value))
    # 将列号 （str）转为列索引值 （int）
    template_key_index = column_index_from_string(template_key)
    template_value_index = column_index_from_string(template_value)
    # 从第line行开始进行vlookup  可根据表头行数进行修改
    for row in range(int(line),ws_template.max_row+1):
        ws_template.cell(row=row, column=template_value_index).value = dic.get(
            ws_template.cell(row=row, column=template_key_index).value)

class VlookupGui(object):
    """
    tkinter GUI界面模板（可选择文件进行操作版）
    """
    def __init__(self):
        """
        创建界面
        """
        # 新建窗口
        self.master = tk.Tk()
        # 在界面顶部添加横幅图片
        self.photo = tk.PhotoImage(file="img\\doge4.gif")
        # self.path 用于存放选择的文件路径
        # self.flag 当程序运行完成后给用户提示信息
        # 注意！这些都是tk.StringVar()对象，不是str，其他地方要用的话要用get()方法获取str
        self.temp_path = tk.StringVar()
        self.source_path = tk.StringVar()
        self.flag = tk.StringVar()
        self.v1 = tk.StringVar()  # 主工作薄工作表序号
        self.v2 = tk.StringVar()   # 主工作薄用于查找的列
        self.v3 = tk.StringVar()   # 主工作薄要填写的列
        self.v4 = tk.StringVar()   # 数据工作薄工作表序号
        self.v5 = tk.StringVar()  # 数据工作薄对应主工作薄的列
        self.v6 = tk.StringVar()  # 数据工作薄提供数据列
        self.v7 = tk.StringVar()  # 从第几行开始vlookup
        # 窗口图片横幅
        self.img_pack()
        # 主界面布局
        self.window()
        # 框架布局
        self.frame()
        # 保持运行
        self.master.mainloop()

    def img_pack(self):
        """
        窗口图片横幅
        :return:
        :rtype:
        """
        # 图片贴上去
        img_lable = tk.Label(self.master, image=self.photo)
        img_lable.pack()

    def window(self):
        """
        主界面布局设置
        :return:
        :rtype:
        """
        # 调整窗口默认大小及在屏幕上的位置
        self.master.geometry("800x900+550+50")
        # 窗口的标题栏，自己修改
        self.master.title("vlookup工具by李加林v2.0")
        # 把标题贴上去，自己修改
        tk.Label(self.master,text="------用于进行VLOOKUP操作------",font=("黑体",20)).pack()
        tk.Label(self.master,text="by李加林",font=("黑体",16)).pack()

    def frame(self):
        """
        框架布局设置
        :return:
        :rtype:
        """
        # 框架贴上去，再在框架里添加Lable，Entry，Button等控件
        frame1 = tk.Frame(self.master)
        frame1.pack()
        frame2 = tk.Frame(self.master)
        frame2.pack()
        # 输入框，标记，按键
        tk.Label(frame1, text="目标路径:", font=("黑体", 16)).grid(row=1, column=0)
        tk.Label(frame1, text="主工作薄->", font=("黑体", 16)).grid(row=2, column=0)
        tk.Entry(frame1, textvariable=self.temp_path, width=50).grid(row=2, column=1)
        tk.Label(frame1, text="数据工作薄->", font=("黑体", 16)).grid(row=3, column=0)
        tk.Entry(frame1, textvariable=self.source_path, width=50).grid(row=3, column=1)
        tk.Button(frame1, text="选择主工作薄", command=self.select_temp_path, font=("黑体", 16)).grid(row=2, column=2)
        tk.Button(frame1, text="选择数据工作薄", command=self.select_source_path, font=("黑体", 16)).grid(row=3, column=2)

        tk.Label(frame2, text="主工作薄工作表序号:", font=("黑体", 16)).grid(row=1, column=0)
        tk.Entry(frame2, textvariable=self.v1).grid(row=1, column=1)
        tk.Label(frame2, text="主工作薄用于查找的列:", font=("黑体", 16)).grid(row=2, column=0)
        tk.Entry(frame2, textvariable=self.v2).grid(row=2, column=1)
        tk.Label(frame2, text="主工作薄要填写的列:", font=("黑体", 16)).grid(row=3, column=0)
        tk.Entry(frame2, textvariable=self.v3).grid(row=3 ,column=1)
        tk.Label(frame2, text="数据工作薄工作表序号:", font=("黑体", 16)).grid(row=4, column=0)
        tk.Entry(frame2, textvariable=self.v4).grid(row=4, column=1)
        tk.Label(frame2, text="数据工作薄对应主工作薄的列:", font=("黑体", 16)).grid(row=5, column=0)
        tk.Entry(frame2, textvariable=self.v5).grid(row=5, column=1)
        tk.Label(frame2, text="数据工作薄提供数据列:", font=("黑体", 16)).grid(row=6, column=0)
        tk.Entry(frame2, textvariable=self.v6).grid(row=6 ,column=1)
        tk.Label(frame2, text="开始vlookup行数:", font=("黑体", 16)).grid(row=7, column=0)
        tk.Entry(frame2, textvariable=self.v7).grid(row=7, column=1)

        # 按这个按钮执行主程序
        tk.Button(frame2, text="开始VLOOKUP", command=self.main, font=("黑体", 16)).grid(row=8, column=1,pady=33)
        tk.Entry(frame2, textvariable=self.flag,state="readonly").grid(row=9, column=1)

    def select_temp_path(self):
        """
        选择文件，并获取文件的绝对路径
        :return:
        :rtype:
        """
        # 选择文件，path_select变量接收文件地址
        # 注意：self.path 是tk.StringVar()对象，而path_select是str变量
        path_select = tkinter.filedialog.askopenfilename()
        # 通过replace函数替换绝对文件地址中的/来使文件可被程序读取
        # 注意：\\转义后为\，所以\\\\转义后为\\
        path_select = path_select.replace("/", "\\\\")
        # self.path设置path_select的值
        self.temp_path.set(path_select)

    def select_source_path(self):
        """
        选择文件，并获取文件的绝对路径
        :return:
        :rtype:
        """
        # 选择文件，path_select变量接收文件地址
        # 注意：self.path 是tk.StringVar()对象，而path_select是str变量
        path_select = tkinter.filedialog.askopenfilename()
        # 通过replace函数替换绝对文件地址中的/来使文件可被程序读取
        # 注意：\\转义后为\，所以\\\\转义后为\\
        path_select = path_select.replace("/", "\\\\")
        # self.path设置path_select的值
        self.source_path.set(path_select)

    def main(self):
        """
        这个是主程序，最好将非gui版的程序封装成类，接收该GUI类的参数后实例化运行
        :return:
        :rtype:
        """
        wb_template = openpyxl.load_workbook(self.temp_path.get())
        wb_source = openpyxl.load_workbook(self.source_path.get())

        ws_template_index = int(self.v1.get()) - 1
        template_key = self.v2.get()
        template_value = self.v3.get()

        ws_source_index = int(self.v4.get()) - 1
        source_key = self.v5.get()
        source_value = self.v6.get()
        line = self.v7.get()
        # 进行vlookup处理
        vlookup(wb_template,ws_template_index,template_key,template_value,wb_source,ws_source_index,source_key,source_value,line)
        wb_template.save(self.temp_path.get().replace('.xlsx','')+'_已进行vlookup.xlsx')
        # 标志设置为处理完成
        self.flag.set("处理完成！")

if __name__ == '__main__':
    # 实例化运行
    app = VlookupGui()