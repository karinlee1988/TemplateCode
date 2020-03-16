#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2020/3/14 16:46
# @Author : karinlee
# @FileName : merge_workbook_gui_openpyxl.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976

import os
import openpyxl
import tkinter as tk
import tkinter.filedialog


def get_xlsx_filename(file_dir):
    """
    构造获取所有需要合并的工作簿路径并生成路径列表的函数
    :param file_dir: 要合并的工作薄所在的文件夹路径
    :type file_dir:  str
    :return: 文件夹下所有.xlsx文件名列表
    :rtype:   list
    """
    xlsx_list = []  # 构造一个用于存放文件名(包括扩展名)的空列表
    for file in os.listdir(file_dir):  # 遍历文件夹file_dir下的所有文件
        if os.path.splitext(file)[1] == '.xlsx':  # 筛选出扩展名是.xlsx的所有文件
            xlsx_list.append(file)  # 将文件扩展名是.xlsx的所有文件的文件名存放到列表list中
    return xlsx_list


def merge_workbooks(path, sheet_index=1, title_row=1):
    """
    合并工作薄
    :param path: 待合并工作薄文件夹
    :type path: str
    :param title_row: 表头行数（从1开始）
    :type title_row: int
    :return: None
    :rtype:  None
    """

    sheet_index = int(sheet_index)  # 转换为int
    title_row = int(title_row)    # 转换为int
    xlsx_filename_list = get_xlsx_filename(path)  # 通过get_xlsx_filename函数获取path路径下所有xlsx文件的文件名
    main_workbook = openpyxl.Workbook()   # 新建合并后的工作薄
    main_worksheet = main_workbook.active
    # 构建表头 表头数据随便从文件列表中第一个文件处获取
    temp_workbook = openpyxl.load_workbook(path + '\\' + xlsx_filename_list[0])
    temp_worksheet = temp_workbook[temp_workbook.sheetnames[sheet_index - 1]]
    list_all_title = []
    list_row_title = []
    for each_title_row in range(1,title_row+1):
        for cell in temp_worksheet[each_title_row]:
            list_row_title.append(cell.value)
        list_all_title.append(list_row_title)
        list_row_title = []
    for title in list_all_title:
        main_worksheet.append(title)
    # 构建数据部分，需要从每个文件中拿取
    for filename in xlsx_filename_list:
        merge_workbook = openpyxl.load_workbook(path + '\\' + filename)
        merge_worksheet = merge_workbook[merge_workbook.sheetnames[sheet_index - 1]]
        list_all = []
        list_row = []
        for row in range(title_row+1, merge_worksheet.max_row + 1):
            for cell in merge_worksheet[row]:  # 对当前行遍历所有单元格
                list_row.append(cell.value)  # list_row是临时的1维列表 在遍历单元格获得每个单元格的值后写入列表 从而存储当前行的数据
            list_all.append(list_row)  # list_all是二维列表 里面的每个元素都是1个list_row
            list_row = []  # 重新初始化list_row列表
        for row in list_all:  # 对于2维列表的每个元素（每个元素就是每个1维列表 这些1维列表就等于excel表中1行的数据
            main_worksheet.append(row)  # 可以直接使用worksheet.append()方法写入工作表中

    main_workbook.save("合并完成.xlsx")


class MergeWorkbook(object):
    """
    用于合并工作薄
    """
    def __init__(self):
        """
        创建界面
        """
        # 新建窗口
        self.master = tk.Tk()
        # 在界面顶部添加横幅图片
        self.photo = tk.PhotoImage(file="img\\doge5.gif")
        # self.path 用于存放选择的文件路径
        # self.flag 当程序运行完成后给用户提示信息
        # self.v1 self.v2接收用户参数 待选用，不够可以加
        # 注意！这些都是tk.StringVar()对象，不是str，其他地方要用的话要用get()方法获取str
        self.path = tk.StringVar()
        self.flag = tk.StringVar()
        self.sheet_index = tk.StringVar()
        self.title_row = tk.StringVar()
        # 窗口图片横幅
        self.img_pack()
        # 主界面布局
        self.window()
        # 框架布局
        self.frame()
        # 保持运行
        self.master.mainloop()

    def img_pack(self):
        """
        窗口图片横幅
        :return:
        :rtype:
        """
        # 图片贴上去
        img_lable = tk.Label(self.master, image=self.photo)
        img_lable.pack()

    def window(self):
        """
        主界面布局设置
        :return:
        :rtype:
        """
        # 调整窗口默认大小及在屏幕上的位置
        self.master.geometry("800x650+600+100")
        # 窗口的标题栏，自己修改
        self.master.title("合并工作薄by李加林v2.0")
        # 把标题贴上去，自己修改
        tk.Label(self.master,text="合并工作薄v2.0",font=("黑体",20)).pack()
        tk.Label(self.master,text="------注意！待合并的工作薄必须是.xlsx格式------",font=("黑体",16)).pack()


    def frame(self):
        """
        框架布局设置
        :return:
        :rtype:
        """
        # 框架贴上去，再在框架里添加Lable，Entry，Button等控件
        frame1 = tk.Frame(self.master)
        frame1.pack()
        # 输入框，标记，按键
        tk.Label(frame1, text="工作薄所在文件夹路径:", font=("黑体", 16)).grid(row=1, column=0)
        tk.Entry(frame1, textvariable=self.path, width=50).grid(row=1, column=1)
        tk.Button(frame1, text="选择文件夹", command=self.select_path, font=("黑体", 16)).grid(row=1, column=2)
        tk.Label(frame1, text="工作表序号：", font=("黑体", 16)).grid(row=2, column=0)
        tk.Entry(frame1, textvariable=self.sheet_index).grid(row=2, column=1)
        tk.Label(frame1, text="表头行数:", font=("黑体", 16)).grid(row=3, column=0)
        tk.Entry(frame1, textvariable=self.title_row).grid(row=3, column=1)
        # 按这个按钮执行主程序
        tk.Button(frame1, text="开始处理", command=self.main, font=("黑体", 16)).grid(row=4, column=1,pady=66)
        tk.Entry(frame1, textvariable=self.flag,state="readonly").grid(row=4, column=2)

    def select_path(self):
        """
        选择文件，并获取文件的绝对路径
        :return:
        :rtype:
        """
        # 选择文件，path_select变量接收文件地址
        # 注意：self.path 是tk.StringVar()对象，而path_select是str变量
        path_select = tkinter.filedialog.askdirectory()
        # 通过replace函数替换绝对文件地址中的/来使文件可被程序读取
        # 注意：\\转义后为\，所以\\\\转义后为\\
        path_select = path_select.replace("/", "\\\\")
        # self.path设置path_select的值
        self.path.set(path_select)

    def main(self):
        """
        这个是主程序，最好将非gui版的程序封装成类，接收该GUI类的参数后实例化运行
        :return:
        :rtype:
        """
        merge_workbooks(path=self.path.get(),sheet_index=self.sheet_index.get(),title_row=self.title_row.get())
        # 标志设置为处理完成
        self.flag.set("处理完成！")

if __name__ == '__main__':
    # path = 'D:\\我的坚果云\\个人学习文档\\python小工具\\合并工作薄\\xlsx文件夹'
    # # print(xlsx_filename_list(path))
    # merge_workbooks(path=path,title_row=2)

    running = MergeWorkbook()

