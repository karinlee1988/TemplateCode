#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2020/1/28 16:56
# @Author : karinlee
# @FileName : workbook_split_for_line_gui_v2.0.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976

"""
    本模块用于按照固定行数将单个工作薄拆分为多个工作薄
    表头默认为1行
"""

import openpyxl
import tkinter as tk
import tkinter.filedialog

class WorkbookSplitRegularLine(object):
    """
    按照固定行数将单个工作薄拆分为多个工作薄
    表头默认为1行
    """

    def __init__(self,full_filename,batche_num):
        """

        :param full_filename: 文件的全文件名（相对路径或绝对路径）
        :param batche_num: 需要拆分的行数
        :type full_filename: str
        :type batche_num: int or str
        """
        self.full_filename = full_filename
        self.batche_num = batche_num
        # 获取self.workbook 要拆分的工作薄对象
        self.workbook = openpyxl.load_workbook(self.full_filename)
        # 获取worksheet，默认为第一个sheet用于拆分
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        # 把生成器转换为列表
        self.lines = list(self.sheet.rows)
        # 获取第一行合并行
        # first_line = lines[0]
        # 获取表头字段行
        self.header = self.lines[0:1]
        # 获取数据行
        self.dataline = self.lines[1:]

    def one_sheet(self,data, sheet_no):
        """
        把一个数据集保存成一个xlsx文件，。
        :param data: 数据集
        :param sheet_no: 拆分文件的流水号
        :type data: list
        :type sheet_no: int
        :return: None
        """
        # 创建一个xlsx文件对象
        wb = openpyxl.Workbook()
        # 取得默认的worksheet
        ws = wb.active
        # ws.title = '新测试表%02d' % (sheet_no+1)   # 设置一个标题
        # # 第一行写入合并行
        # ws.cell(row=1, column=1).value = first_line[0].value
        # # 该行所有列合并
        # ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(first_line))
        # 写头：循环写每个字段的值：行从1开始，所以表头行索引是2
        header_idx = 0
        for col in self.header[0]:
            ws.cell(row=1, column=(header_idx + 1)).value = col.value
            header_idx += 1
        # 纪录行索引
        row_idx = 0
        for row_ in data:
            col_idx = 0
            for col in row_:
                # 数据行行从2开始
                ws.cell(row=(row_idx+2) ,column=(col_idx + 1)).value = col.value
                # 纪录列索引
                col_idx += 1
            row_idx += 1
        # 保存文件
        wb.save(self.full_filename+'-%02d.xlsx' % (sheet_no + 1))

    def dealing(self):
        # 计算拆分文件个数
        batches = (len(self.dataline) // self.batche_num) + 1
        # 循环写数据集到每一个文件
        for pt in range(batches):
            # 取数据集，每个数据集最多3000行，最后一个不足3000行，直接处理。
            lines_ = self.dataline[pt * self.batche_num : (pt + 1) * self.batche_num]
            self.one_sheet(lines_,pt)

class WorkbookSplitRegularLineGUI(object):
    """
    tkinter GUI界面模板（可选择文件进行操作）
    """
    def __init__(self):
        """
        创建界面
        """
        # 新建窗口
        self.master = tk.Tk()
        self.photo = tk.PhotoImage(file="img\\si2.gif")
        # self.path 用于存放选择的文件路径,self.v1 self.v2接收用户参数
        # 注意！这些都是tk.StringVar()对象，不是str，其他地方要用的话要用get()方法获取str
        self.path = tk.StringVar()
        self.flag = tk.StringVar()
        self.line = tk.StringVar()

        # 窗口图片横幅
        self.img_pack()
        # 主界面布局
        self.window()
        # 框架布局
        self.frame()
        # 保持运行
        self.master.mainloop()

    def img_pack(self):
        """
        窗口图片横幅
        :return: None
        :rtype:  None
        """
        # 图片贴上去
        img_lable = tk.Label(self.master, image=self.photo)
        img_lable.pack()

    def window(self):
        """
        主界面布局设置
        :return:
        :rtype:
        """
        self.master.geometry("800x600+600+100")
        self.master.title("固定行数拆分excel工作薄工具by李加林v2.0")
        # 把标题贴上去
        tk.Label(self.master, text="固定行数拆分xlsx单个工作薄为多个工作薄\n", font=("黑体", 20)).pack()
        tk.Label(self.master, text="表头默认为1行\n\n", font=("黑体", 16)).pack()

    def frame(self):
        """
        框架布局设置
        :return: None
        :rtype: None
        """
        # 框架贴上去，再在框架里添加Lable，Entry，Button等控件
        frame1 = tk.Frame(self.master)
        frame1.pack()
        # 输入框，标记，按键
        tk.Label(frame1, text="目标路径:", font=("黑体", 16)).grid(row=1, column=0)
        tk.Entry(frame1, textvariable=self.path, width=50).grid(row=1, column=1)
        tk.Button(frame1, text="路径选择", command=self.select_path, font=("黑体", 16)).grid(row=1, column=2)
        tk.Label(frame1, text="请输入固定拆分行数：", font=("黑体", 16)).grid(row=2, column=0)
        tk.Entry(frame1, textvariable=self.line).grid(row=2, column=1)
        # 按这个按钮执行主程序
        tk.Button(frame1, text="开始处理", command=self.main, font=("黑体", 16)).grid(row=4, column=1, pady=66)
        tk.Entry(frame1, textvariable=self.flag, state="readonly",width=20).grid(row=4, column=2, pady=66)

    def select_path(self):
        """
        选择文件，并获取文件的绝对路径
        :return: None
        :rtype:  None
        """
        # 选择文件，path_select变量接收文件地址
        path_select = tkinter.filedialog.askopenfilename()
        # 通过replace函数替换绝对文件地址中的/来使文件可被程序读取
        # 注意：\\转义后为\，所以\\\\转义后为\\
        path_select = path_select.replace("/", "\\\\")
        # self.path设置path_select的值
        self.path.set(path_select)

    def main(self):
        """
        这个是主程序
        :return: None
        :rtype:  None
        """
        run = WorkbookSplitRegularLine(self.path.get(),int(self.line.get()))
        run.dealing()
        # 标志设置为处理完成
        self.flag.set("处理完成！")

if __name__ == '__main__':
    # 实例化gui开始运行
    app = WorkbookSplitRegularLineGUI()