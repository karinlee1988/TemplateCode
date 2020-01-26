#!/usr/bin/env python3
# -*- coding: utf-8 -*-


from openpyxl import Workbook
from openpyxl import load_workbook
filename = '示例.xlsx'
# filename = '用户数据v1'

workbook = load_workbook(filename)
# 得到所有worksheet

# 获取第一个worksheet
sheet = workbook[workbook.sheetnames[0]]
# 把生成器转换为列表
lines = list(sheet.rows)

# 获取第一行合并行
# first_line = lines[0]
# 获取表头字段行
header = lines[0:1]
# 获取数据行
lines = lines[1:]


# 把一个数据集保存成一个xlsx文件，参数1：数据集，参数2：拆分文件的流水号。
def one_sheet(lines_1000, sheet_no):
    # 创建一个xlsx文件对象
    wb = Workbook()
    ws = wb.active    # 取得默认的worksheet
    # ws.title = '新测试表%02d' % (sheet_no+1)   # 设置一个标题
    # # 第一行写入合并行
    # ws.cell(row=1, column=1).value = first_line[0].value
    # # 该行所有列合并
    # ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(first_line))
    # 写头：循环写每个字段的值：行从1开始，所以表头行索引是2
    header_idx = 0
    for col in header[0]:
        ws.cell(row=1, column=(header_idx + 1)).value = col.value
        # print(col.value)
        header_idx += 1

    row_idx = 0   # 纪录行索引
    for row_ in lines_1000:
        col_idx = 0
        for col in row_:
            ws.cell(row=(row_idx+2) ,column=(col_idx + 1)).value = col.value    # 数据行行从2开始
            col_idx += 1    # 纪录列索引
        row_idx += 1

    # 保存文件
    wb.save(filename+'-%02d.xlsx' % (sheet_no + 1))

# 每个文件数据行数
batche_num = 3000
# 计算拆分文件个数
batches = (len(lines) // batche_num) + 1
# 循环写数据集到每一个文件
for pt in range(batches):
    # 取数据集，每个数据集最多3000行，最后一个不足3000行，直接处理。
    lines_ = lines[pt * batche_num : (pt + 1) * batche_num]
    one_sheet(lines_,pt)