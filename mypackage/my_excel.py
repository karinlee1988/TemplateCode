#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2020/2/11 14:54
# @Author : karinlee
# @FileName : my_excel.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976

"""
本模块包含相关自建函数，用于处理excel表格。
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

def vlookup(
        wb_template,
        ws_template_index,
        template_key,
        template_value,
        wb_source,
        ws_source_index,
        source_key,
        source_value,
        line
        ):
    """
    对2个不同的工作薄执行vlookup操作

    模板工作薄：需要写入数据的工作薄
    数据工作薄：根据模板工作薄提供的条件（列）在数据工作薄中查找，提供数据来源的工作薄
    注意！函数执行完后，只对wb_template对象进行了数据写入。在函数外部还需wb_template.save("filename.xlsx"),vlookup后的数据才能保存为excel表。

    :param:
        'wb_template': 模板工作薄对象
        'ws_template_index': 需要处理的模板工作表索引号
        'template_key': 模板工作表key所在列号
        'template_value': 模板工作表value需要填写的列号
        'wb_source': 数据工作薄对象
        'ws_source_index': 需要处理的数据工作表索引号
        'source_key':  数据工作表key所在列号
        'source_value': 数据工作表value所在列号
        'line' :从第几行开始vlookup

    :type:
        'wb_template': class workbook
        'ws_template_index': int
        'template_key': str
        'template_value': str
        'wb_source': class workbook
        'ws_source_index': int
        'source_key':  str
        'source_value': str
        'line' :int

    :return: None

    """
    # 获取数据工作表对象
    ws_source = wb_source[wb_source.sheetnames[ws_source_index]]
    # 获取模板工作表对象
    ws_template = wb_template[wb_template.sheetnames[ws_template_index]]
    # 获取数据工作表的查找列和数据列，分别生成2个元组
    source_key_tuple = ws_source[source_key]
    source_value_tuple = ws_source[source_value]
    # 创建2个列表，遍历元组，将元组中每个单元格的值添加到列表中
    list_key =[]
    list_value = []
    for cell in source_key_tuple:
        list_key.append(cell.value)
    for cell in source_value_tuple:
        list_value.append(cell.value)
    # 通过数据工作表的key列和value列，创建好需要进行vlookup的字典
    dic = dict(zip(list_key,list_value))
    # 将列号 （str）转为列索引值 （int）
    template_key_index = column_index_from_string(template_key)
    template_value_index = column_index_from_string(template_value)
    # 从第line行开始进行vlookup  可根据表头行数进行修改
    for row in range(int(line),ws_template.max_row+1):
        # try:
        #     ws_template.cell(row=row,column=template_value_index).value = dic[
        #     ws_template.cell(row=row,column=template_key_index).value]
        # except KeyError:
        #     #找不到数据 相应的单元格填上#N/A
        #     ws_template.cell(row=row, column=template_value_index).value = "#N/A"

        #采用dict.get()避免出现keyerror
        ws_template.cell(row=row, column=template_value_index).value = dic.get(
            ws_template.cell(row=row, column=template_key_index).value)

def get_xlsx_full_filename(folder_path):
    """
    获取待处理文件夹里所有后缀为.xlsx的全文件名

    :param folder_path : 文件夹路径
    :type folder_path : str

    :return 文件夹里所有后缀为.xlsx的全文件名列表
    :rtype list
    """
    filename_list = os.listdir(folder_path)
    xlsx_list = []
    for filename in filename_list:
        # os.path.splitext():分离文件名与扩展名
        if os.path.splitext(filename)[1] == '.xlsx':
            xlsx_list.append(folder_path+filename)
    return xlsx_list

def get_xlsx_filename(folder_path):
    """
    获取待处理文件夹里所有后缀为.xlsx的全文件名

    :param folder_path : 文件夹路径
    :type folder_path : str

    :return 文件夹里所有后缀为.xlsx的文件名列表
    :rtype list
    """
    filename_list = os.listdir(folder_path)
    xlsx_list = []
    for filename in filename_list:
        # os.path.splitext():分离文件名与扩展名
        if os.path.splitext(filename)[1] == '.xlsx':
            xlsx_list.append(filename)
    return xlsx_list

def worksheet_save_as(path,workbook):
    """
    将一个工作薄里面的多个工作表分别另存为独立的工作薄，独立的工作薄名称为原工作薄各工作表表名

    :param path :另存为的路径
    :type path : str

    :param workbook:需要进行工作表另存为的workbook对象
    :type workbook: class workbook

    :return: None

    """
    sheetname_list = workbook.sheetnames
    for name in sheetname_list:
        worksheet = workbook[name]
        # 创建新的Excel
        workbook_new = openpyxl.Workbook()
        # 获取当前sheet
        worksheet_new = workbook_new.active
        # 两个for循环遍历整个excel的单元格内容
        for i, row in enumerate(worksheet.iter_rows(),start=1): #enumerate()从1开始  # 或者for i, row in enumerate(worksheet.rows):
            for j, cell in enumerate(row,start=1):#enumerate()从1开始
                # 写入新Excel
                worksheet_new.cell(row=i, column=j, value=cell.value)
                # 设置新Sheet的名称
                worksheet_new.title = name
        workbook_new.save(path + name + '.xlsx')
