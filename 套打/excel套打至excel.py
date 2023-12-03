#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2023/12/03 14:46
# @Author : karinlee
# @FileName : excel套打至excel.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @Personal website : https://karinlee.cn/

"""
用于套打用。将一个excel数据表格多行的内容分别套打到excel 模板中，excel数据表格每行生成一份套打的excel模板文档
"""

import os
import time
import openpyxl


class ExcelPrintToExcel(object):
    def __init__(self, template_path: str, source_path: str, isprint=False):
        """
        20231203 test OK


        :param template_path: 最终需要打印的模板表格
        :type template_path:str

        :param source_path: 数据来源的表格
        :type source_path:str

        :param isprint: 是否需要实际操作打印机执行打印命令，默认为否
        :type isprint: bool

        """

        self.source_path = source_path
        self.template_path = template_path
        self.isprint = isprint
        # 获取excel文档工作薄、工作表对象
        self.workbook_template = openpyxl.load_workbook(self.template_path)
        self.workbook_source = openpyxl.load_workbook(self.source_path)
        self.worksheet_template = self.workbook_template.active
        self.worksheet_source = self.workbook_source.active

    def main(self):
        """
            程序入口，可以根据实际情况修改后进行套打
        """
        # -------------------
        # 这里修改套打数据（数据来源的表格）的开始编号->结束编号
        # 编号也等于序号，用于套打时能方便定位寻找相应文档。一般来说，excel表第一行是表头，从第二行开始为编号（序号）1的数据，
        # 第三行为编号（序号）2的数据...
        tag_begin = 1
        tag_end = 10
        # -------------------
        # -------------------
        # 这里修改套打时要替换的元素内容，这里准备了5个元素，可以根据实际的表格结构增加或删除元素
        for row in range(tag_begin + 1, tag_end + 2):
            tag = str(self.worksheet_source.cell(row=row, column=1).value)
            company = str(self.worksheet_source.cell(row=row, column=2).value)
            company_num = str(self.worksheet_source.cell(row=row, column=3).value)
            name = str(self.worksheet_source.cell(row=row, column=4).value)
            sfz = str(self.worksheet_source.cell(row=row, column=5).value)

            # 开始写入模板
            self.worksheet_template['D4'].value = name
            self.worksheet_template['D5'].value = sfz
            self.worksheet_template['D6'].value = company_num
            self.worksheet_template['N6'].value = company
            self.worksheet_template['A1'].value = tag
            # 保存(另存为)
            self.workbook_template.save(tag + name + '.xlsx')
            print(f">>> 序号：{tag},姓名： {name} ，身份证号码：{sfz} 已处理完成！<<<")
            # -------------------
            # 是否需要实际打印，默认为否
            if self.isprint:
                os.startfile(tag + name + '.xlsx', 'print')
                # 根据打印机的情况设置延时
                time.sleep(10)


if __name__ == '__main__':
    p = ExcelPrintToExcel('tests\\最终打印表格_测试excel套打至excel.xlsx', 'tests\\数据来源_测试excel套打至excel.xlsx',
                          isprint=False)
    p.main()
